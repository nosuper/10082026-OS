"""Turning a bank's spreadsheet into rows (#150).

The thin half of statement import: everything that needs `openpyxl` and
a file on disk lives here, and everything that needs judgement lives in
`auraos.lib.statement`, which takes rows and knows nothing about
spreadsheets. So the arithmetic, the references and the matcher are
testable with no file at all, and this module has one job - find the
blocks and hand over their cells.

**Blocks are found by their labels, not by their row numbers.** The
sample's summary sits on row 7 and its table starts on row 9, and a bank
that adds a line to its header would move both. Labels are what the bank
promises; row numbers are what it happened to print this month.

**A cell is handed over as it arrived.** `openpyxl` resolves what it can
- dates come back as datetimes, amounts as floats - and the rest stays
text, which is why `lib.statement.to_amount` accepts both. Converting
here would put two parsers in the app.
"""

from __future__ import annotations

import re
from typing import Any, Iterable

# What the bank calls its four summary figures, and what we call them.
SUMMARY_LABELS = {
    "số dư đầu": "opening",
    "(-) tổng tiền rút ra": "withdrawn",
    "(+) tổng tiền gửi vào": "deposited",
    "số dư cuối": "closing",
}

# The transaction table's own headings.
COLUMN_LABELS = {
    "ngày hiệu lực": "effective_on",
    "ngày giao dịch": "transacted_at",
    "số gd": "sequence",
    "nội dung giao dịch": "description",
    "số tiền rút ra": "withdrawn",
    "số tiền gửi vào": "deposited",
    "số dư": "running_balance",
}

ACCOUNT_IN_HEADER = re.compile(r"S[ốo]\s*t[àa]i\s*kho[ảa]n[^:]*:\s*(\d+)", re.IGNORECASE)
PERIOD_IN_HEADER = re.compile(
    r"T[ừu]\s*ng[àa]y:\s*(\d{2}/\d{2}/\d{4}).*?[ĐD][ếe]n\s*ng[àa]y:\s*(\d{2}/\d{2}/\d{4})",
    re.IGNORECASE | re.DOTALL,
)


def _text(value: Any) -> str:
    return " ".join(str(value or "").split()).lower()


def sheet_rows(path: str) -> list[list[Any]]:
    """Every row of the first sheet, as lists of cell values.

    Imported inside the function because `openpyxl` is Frappe's
    dependency rather than ours: an app that fails to import when a
    statement is not being read would be a poor trade for one line.
    """
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        book.close()


def read(rows: list[list[Any]]) -> dict:
    """The whole statement: its header, its four totals, its transactions.

    Raises `ValueError` naming what was missing, because every caller of
    this is a person who has just chosen a file and can choose another
    one - "not a statement we recognise" is a better answer than a
    traceback about a None.
    """
    header = "\n".join(str(cell) for row in rows[:8] for cell in row if cell)
    account = ACCOUNT_IN_HEADER.search(header)
    period = PERIOD_IN_HEADER.search(header)

    summary_at = None
    columns_at = None
    for index, row in enumerate(rows):
        labels = {_text(cell) for cell in row}
        if summary_at is None and SUMMARY_LABELS.keys() <= labels:
            summary_at = index
        if columns_at is None and "ngày hiệu lực" in labels:
            columns_at = index
    if summary_at is None:
        raise ValueError("no summary block: this file does not look like a statement")
    if columns_at is None:
        raise ValueError("no transaction table: this file does not look like a statement")

    heads = rows[summary_at]
    figures = rows[summary_at + 1] if summary_at + 1 < len(rows) else []
    summary = {}
    for column, cell in enumerate(heads):
        field = SUMMARY_LABELS.get(_text(cell))
        if field and column < len(figures):
            summary[field] = figures[column]

    columns = {}
    for column, cell in enumerate(rows[columns_at]):
        field = COLUMN_LABELS.get(_text(cell))
        if field:
            columns[field] = column

    lines = []
    for row in rows[columns_at + 1 :]:
        cells = {field: row[column] if column < len(row) else None
                 for field, column in columns.items()}
        # The table ends where the sequence numbers do. A bank that adds
        # a footer would otherwise contribute a row of Nones that fails
        # the arithmetic check for the wrong reason.
        if not str(cells.get("sequence") or "").strip():
            continue
        lines.append(cells)

    return {
        "account_number": account.group(1) if account else None,
        "period_from": period.group(1) if period else None,
        "period_to": period.group(2) if period else None,
        "summary": summary,
        "lines": lines,
    }


def read_file(path: str) -> dict:
    """`read`, from a path on disk."""
    return read(sheet_rows(path))
